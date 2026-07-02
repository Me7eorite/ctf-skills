import json
import tarfile
import tempfile
import unittest
import zipfile
from datetime import date
from pathlib import Path
from unittest.mock import patch

from openpyxl import load_workbook

from core.paths import ProjectPaths
from packing import (
    IMAGE_HEADERS,
    OVERVIEW_HEADERS,
    Packer,
    PackerOptions,
    PackingError,
)


class PackingTests(unittest.TestCase):
    def setUp(self):
        self.temp = tempfile.TemporaryDirectory()
        self.addCleanup(self.temp.cleanup)
        self.paths = ProjectPaths(
            root=Path(self.temp.name) / "factory",
            repository=Path(self.temp.name),
        )
        self.paths.initialize()
        self.output = self.paths.work / "bundle"

    def _challenge(
        self,
        challenge_id: str,
        category: str,
        *,
        build_status: str = "passed",
        solve_status: str = "passed",
        with_deploy: bool = False,
        with_attachment: bool = True,
    ) -> Path:
        challenge = self.paths.challenges / category / f"{challenge_id}-demo"
        (challenge / "writenup").mkdir(parents=True)
        (challenge / "writenup" / "wp.md").write_text(
            "# 题目分析\n\n这是中文题解。\n\n```python\nprint('flag')\n```\n",
            encoding="utf-8",
        )
        (challenge / "writenup" / "exp.py").write_text(
            "print('flag{demo}')\n", encoding="utf-8"
        )
        if with_attachment:
            (challenge / "dist").mkdir()
            (challenge / "dist" / "checker.bin").write_bytes(b"artifact")
        if with_deploy:
            (challenge / "deploy" / "src").mkdir(parents=True)
            (challenge / "deploy" / "src" / "app.py").write_text(
                "print('ok')\n", encoding="utf-8"
            )
            (challenge / "deploy" / "Dockerfile").write_text(
                "FROM python:3.13-slim\n", encoding="utf-8"
            )
        metadata = {
            "id": challenge_id,
            "title": "Demo Challenge",
            "description": "A synthetic challenge",
            "category": category,
            "difficulty": "medium",
            "points": 200,
            "primary_technique": "testing",
            "build_status": build_status,
            "solve_status": solve_status,
            "flag": "flag{demo}",
            "port": 8080 if category == "web" else 9001,
        }
        (challenge / "metadata.json").write_text(
            json.dumps(metadata), encoding="utf-8"
        )
        return challenge

    def _pack(self, **options):
        return Packer(
            self.paths,
            PackerOptions(
                skip_docker=True,
                generated_on=date(2026, 6, 9),
                **options,
            ),
        ).pack(self.output)

    def test_reverse_bundle_uses_delivery_prefix_and_normalizes_solver(self):
        self._challenge("re-0001", "re")

        summary = self._pack()

        self.assertEqual(summary["challenges"], 1)
        tools = self.output / "工具" / "js-reverse-re-0001exp.zip"
        enclosure = (
            self.output
            / "题库资源"
            / "deploy"
            / "enclosure"
            / "js-reverse-re-0001.zip"
        )
        with zipfile.ZipFile(tools) as archive:
            self.assertEqual(set(archive.namelist()), {"wp.md", "exp.py"})
        with zipfile.ZipFile(enclosure) as archive:
            self.assertEqual(archive.namelist(), ["checker.bin"])

        pdf = (
            self.output
            / "题库资源"
            / "deploy"
            / "report"
            / "js-reverse-re-0001.pdf"
        )
        self.assertTrue(pdf.read_bytes().startswith(b"%PDF"))

    def test_web_emits_deploy_tree_without_enclosure(self):
        self._challenge("web-0001", "web", with_deploy=True, with_attachment=False)

        self._pack()

        deployment = self.output / "题库资源" / "deploy" / "js-web-web-0001.zip"
        with zipfile.ZipFile(deployment) as archive:
            self.assertEqual(
                set(archive.namelist()),
                {"deploy/Dockerfile", "deploy/src/app.py"},
            )
        enclosure = self.output / "题库资源" / "deploy" / "enclosure"
        self.assertEqual(list(enclosure.iterdir()), [])

    def test_containerized_challenge_rejects_port_zero(self):
        challenge = self._challenge(
            "web-0001", "web", with_deploy=True, with_attachment=False
        )
        metadata_path = challenge / "metadata.json"
        metadata = json.loads(metadata_path.read_text(encoding="utf-8"))
        metadata["port"] = 0
        metadata_path.write_text(json.dumps(metadata), encoding="utf-8")

        with self.assertRaisesRegex(PackingError, "invalid port"):
            self._pack()

    def test_pwn_enclosure_is_opt_in(self):
        self._challenge("pwn-0001", "pwn", with_deploy=True)
        self._pack()
        enclosure = (
            self.output
            / "题库资源"
            / "deploy"
            / "enclosure"
            / "js-pwn-pwn-0001.zip"
        )
        self.assertFalse(enclosure.exists())

        self._pack(include_pwn_attachments=True)

        self.assertTrue(enclosure.exists())

    def test_workbooks_have_required_headers_and_only_passed_challenges(self):
        self._challenge("re-0001", "re")
        self._challenge("re-0002", "re", build_status="failed")

        self._pack()

        overview_book = load_workbook(
            self.output / "题库资源" / "ctf-overview.xlsx", read_only=True
        )
        self.addCleanup(overview_book.close)
        overview = overview_book.active
        self.assertEqual(
            [cell.value for cell in next(overview.iter_rows())], OVERVIEW_HEADERS
        )
        self.assertEqual(overview.max_row, 2)
        self.assertEqual(overview["A2"].value, "re-0001")

        image_book = load_workbook(
            self.output / "虚拟机资源" / "镜像模板.xlsx", read_only=True
        )
        self.addCleanup(image_book.close)
        images = image_book.active
        self.assertEqual(
            [cell.value for cell in next(images.iter_rows())], IMAGE_HEADERS
        )
        self.assertEqual(images.max_row, 1)

    def test_pack_can_scope_to_single_challenge_id(self):
        self._challenge("re-0001", "re")
        self._challenge("re-0002", "re")

        summary = Packer(
            self.paths,
            PackerOptions(
                skip_docker=True,
                generated_on=date(2026, 6, 9),
                include_ids={"re-0002"},
            ),
        ).pack(self.output)

        self.assertEqual(summary["challenges"], 1)
        overview_book = load_workbook(
            self.output / "题库资源" / "ctf-overview.xlsx", read_only=True
        )
        self.addCleanup(overview_book.close)
        overview = overview_book.active
        self.assertEqual(overview.max_row, 2)
        self.assertEqual(overview["A2"].value, "re-0002")

    def test_pack_replaces_stale_output(self):
        stale = self.output / "工具" / "old.txt"
        stale.parent.mkdir(parents=True)
        stale.write_text("old", encoding="utf-8")
        unrelated = self.output / "operator-note.txt"
        unrelated.write_text("keep", encoding="utf-8")
        self._challenge("re-0001", "re")

        self._pack()

        self.assertFalse(stale.exists())
        self.assertTrue(unrelated.exists())

    def test_unsolved_challenge_is_excluded_from_pack(self):
        # build_status=passed but solve_status not passed must NOT ship:
        # a broken/pending/hardcoded solver is a non-publishable challenge.
        self._challenge("re-0001", "re")
        self._challenge("re-0002", "re", solve_status="pending")
        self._challenge("re-0003", "re", solve_status="failed")

        self._pack()

        overview_book = load_workbook(
            self.output / "题库资源" / "ctf-overview.xlsx", read_only=True
        )
        self.addCleanup(overview_book.close)
        overview = overview_book.active
        self.assertEqual(overview.max_row, 2)  # header + only re-0001
        self.assertEqual(overview["A2"].value, "re-0001")

    def test_non_chinese_writeup_produces_warning(self):
        challenge = self._challenge("re-0001", "re")
        (challenge / "writenup" / "wp.md").write_text(
            "# Analysis\n\nEnglish only.\n", encoding="utf-8"
        )

        summary = self._pack()

        self.assertEqual(len(summary["warnings"]), 1)
        self.assertIn("contains no CJK text", summary["warnings"][0])

    def test_missing_docker_warns_but_keeps_non_tar_outputs(self):
        self._challenge("web-0001", "web", with_deploy=True, with_attachment=False)
        packer = Packer(
            self.paths,
            PackerOptions(generated_on=date(2026, 6, 9)),
        )

        with patch("packing.packer.shutil.which", return_value=None):
            summary = packer.pack(self.output)

        self.assertIn("docker CLI unavailable", summary["warnings"][0])
        self.assertTrue(
            (self.output / "题库资源" / "deploy" / "js-web-web-0001.zip").exists()
        )
        self.assertEqual(list((self.output / "虚拟机资源" / "docker-tar").iterdir()), [])

    def test_require_docker_fails_when_cli_is_missing(self):
        self._challenge("web-0001", "web", with_deploy=True, with_attachment=False)
        packer = Packer(
            self.paths,
            PackerOptions(require_docker=True, generated_on=date(2026, 6, 9)),
        )

        with patch("packing.packer.shutil.which", return_value=None):
            with self.assertRaisesRegex(PackingError, "docker CLI unavailable"):
                packer.pack(self.output)


class SaveDockerTests(unittest.TestCase):
    @staticmethod
    def _completed(returncode=0, stdout="", stderr=""):
        return type(
            "P", (), {"returncode": returncode, "stdout": stdout, "stderr": stderr}
        )()

    @staticmethod
    def _image_tar(tar_path: Path, repo_tags) -> None:
        """写一个最小的 docker 镜像存档（含 manifest.json）。"""
        import io

        manifest = json.dumps([{"RepoTags": repo_tags}]).encode("utf-8")
        with tarfile.open(tar_path, "w") as archive:
            info = tarfile.TarInfo("manifest.json")
            info.size = len(manifest)
            archive.addfile(info, io.BytesIO(manifest))

    def test_saves_then_prunes_self_generated_image(self):
        from packing.docker import _save_docker

        with tempfile.TemporaryDirectory() as tmp:
            output = Path(tmp)
            errors: list = []
            warnings: list = []
            metadata = {
                "id": "web-0001",
                "port": 8080,
                "docker_image": "web-0001-sqli:latest",
            }
            expected = output / "web-0001-sqli[8080]-20260628.tar"

            def fake_run(cmd, *args, **kwargs):
                # docker save 这一步真正写出存档，供后续校验读取。
                if cmd[:2] == ["docker", "save"]:
                    self._image_tar(expected, ["web-0001-sqli:latest"])
                return self._completed()

            with patch(
                "packing.docker.subprocess.run", side_effect=fake_run
            ) as run:
                tar_path, image_row = _save_docker(
                    "docker",
                    metadata,
                    "web-0001-sqli",
                    output,
                    date(2026, 6, 28),
                    errors,
                    warnings,
                    require_docker=True,
                )

        # 文件名规则恢复原样：题目名[端口]-YYYYMMDD.tar。
        self.assertEqual(tar_path, expected)
        self.assertEqual(image_row[1], "web-0001-sqli[8080]-20260628.tar")
        self.assertEqual(errors, [])
        self.assertEqual(warnings, [])
        # 顺序：inspect → save → 校验通过后 image rm。
        commands = [call.args[0] for call in run.call_args_list]
        self.assertEqual(
            commands,
            [
                ["docker", "image", "inspect", "web-0001-sqli:latest"],
                ["docker", "save", "-o", str(expected), "web-0001-sqli:latest"],
                ["docker", "image", "rm", "web-0001-sqli:latest"],
            ],
        )

    def test_keeps_image_when_save_verification_fails(self):
        from packing.docker import _save_docker

        with tempfile.TemporaryDirectory() as tmp:
            output = Path(tmp)
            errors: list = []
            warnings: list = []
            metadata = {
                "id": "web-0001",
                "port": 8080,
                "docker_image": "web-0001-sqli:latest",
            }

            # inspect ok、save 返回 0 但没写出有效 tar（校验必失败）。
            with patch(
                "packing.docker.subprocess.run",
                side_effect=[self._completed(), self._completed()],
            ) as run:
                with self.assertRaisesRegex(PackingError, "verification failed"):
                    _save_docker(
                        "docker",
                        metadata,
                        "web-0001-sqli",
                        output,
                        date(2026, 6, 28),
                        errors,
                        warnings,
                        require_docker=True,
                    )
            # 校验未过：只调用了 inspect、save，绝不能执行 image rm。
            commands = [call.args[0] for call in run.call_args_list]
            self.assertEqual([cmd[:2] for cmd in commands], [["docker", "image"], ["docker", "save"]])
            self.assertNotIn("rm", [cmd[-1] for cmd in commands])

    def test_prune_failure_warns_but_keeps_tar(self):
        from packing.docker import _save_docker

        with tempfile.TemporaryDirectory() as tmp:
            output = Path(tmp)
            errors: list = []
            warnings: list = []
            metadata = {
                "id": "web-0001",
                "port": 8080,
                "docker_image": "web-0001-sqli:latest",
            }
            expected = output / "web-0001-sqli[8080]-20260628.tar"

            def fake_run(cmd, *args, **kwargs):
                if cmd[:2] == ["docker", "save"]:
                    self._image_tar(expected, ["web-0001-sqli:latest"])
                    return self._completed()
                if cmd[:3] == ["docker", "image", "rm"]:
                    return self._completed(returncode=1, stderr="image is in use")
                return self._completed()

            with patch("packing.docker.subprocess.run", side_effect=fake_run):
                tar_path, image_row = _save_docker(
                    "docker",
                    metadata,
                    "web-0001-sqli",
                    output,
                    date(2026, 6, 28),
                    errors,
                    warnings,
                    require_docker=True,
                )

        # 删除失败不致命：tar 仍交付，仅产生告警。
        self.assertEqual(tar_path, expected)
        self.assertEqual(errors, [])
        self.assertEqual(len(warnings), 1)
        self.assertIn("image not pruned", warnings[0])

    def test_refuses_image_belonging_to_another_challenge(self):
        from packing.docker import _save_docker

        with tempfile.TemporaryDirectory() as tmp:
            output = Path(tmp)
            errors: list = []
            warnings: list = []
            metadata = {
                "id": "web-0001",
                "port": 8080,
                # 偏移：指向了别的题/基础镜像。
                "docker_image": "web-0009-other:latest",
            }
            with patch("packing.docker.subprocess.run") as run:
                with self.assertRaisesRegex(PackingError, "does not belong"):
                    _save_docker(
                        "docker",
                        metadata,
                        "web-0001-sqli",
                        output,
                        date(2026, 6, 28),
                        errors,
                        warnings,
                        require_docker=True,
                    )
            # 既然归属校验未过，绝不能调用 docker。
            run.assert_not_called()

    def test_errors_when_self_image_missing_on_host(self):
        from packing.docker import _save_docker

        with tempfile.TemporaryDirectory() as tmp:
            output = Path(tmp)
            errors: list = []
            warnings: list = []
            metadata = {
                "id": "web-0001",
                "port": 8080,
                "docker_image": "web-0001-sqli:latest",
            }
            # image inspect 返回非零：镜像不存在。
            with patch(
                "packing.docker.subprocess.run",
                side_effect=[self._completed(returncode=1, stderr="No such image")],
            ):
                with self.assertRaisesRegex(PackingError, "not present on host"):
                    _save_docker(
                        "docker",
                        metadata,
                        "web-0001-sqli",
                        output,
                        date(2026, 6, 28),
                        errors,
                        warnings,
                        require_docker=True,
                    )

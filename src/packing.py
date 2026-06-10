"""Render generated challenges into the delivery format v2 bundle."""

from __future__ import annotations

import re
import shutil
import subprocess
import zipfile
from dataclasses import dataclass
from datetime import date
from pathlib import Path
from typing import Any, Iterable

from jsonio import read_json
from paths import ProjectPaths

CATEGORY_PREFIXES = {
    "crypto": "crypto",
    "web": "web",
    "pwn": "pwn",
    "re": "reverse",
    "reverse": "reverse",
    "misc": "misc",
    "stego": "stego",
    "forensics": "forensics",
    "ics": "ics",
    "ai": "ai",
    "cloud": "cloud",
    "mobile": "mobile",
    "blockchain": "blockchain",
    "iot": "iot",
    "auto": "auto",
    "data": "data",
    "malware": "malware",
    "osint": "osint",
}

CATEGORY_LABELS = {
    "crypto": "Crypto",
    "web": "Web",
    "pwn": "Pwn",
    "re": "Reverse",
    "reverse": "Reverse",
    "misc": "Misc",
    "stego": "Stego",
    "forensics": "Forensics",
    "ics": "ICS",
    "ai": "AI",
    "cloud": "Cloud",
    "mobile": "Mobile",
    "blockchain": "Blockchain",
    "iot": "IoT",
    "auto": "Auto",
    "data": "Data",
    "malware": "Malware",
    "osint": "OSINT",
}

ENCLOSURE_RULES = {
    "web": "skip",
    "pwn": "optional",
    "cloud": "optional",
}

OVERVIEW_HEADERS = [
    "题目ID",
    "题目名称",
    "题目描述",
    "题型",
    "难度",
    "考点",
    "分值",
    "flag格式",
    "状态",
]

IMAGE_HEADERS = ["题目名称", "镜像文件", "端口", "基础镜像", "启动命令"]

class PackingError(RuntimeError):
    """Raised when a required delivery artifact cannot be produced."""


@dataclass(frozen=True)
class PackerOptions:
    include_pwn_attachments: bool = False
    skip_docker: bool = False
    require_docker: bool = False
    generated_on: date | None = None


class Packer:
    def __init__(self, paths: ProjectPaths, options: PackerOptions | None = None):
        self.paths = paths
        self.options = options or PackerOptions()
        self.warnings: list[str] = []
        self.errors: list[str] = []

    def pack(self, out_dir: Path | None = None) -> dict[str, Any]:
        self.warnings.clear()
        self.errors.clear()
        if self.options.skip_docker and self.options.require_docker:
            raise PackingError("--skip-docker and --require-docker cannot be combined")

        output = (out_dir or self.paths.delivery_bundle).resolve()
        self._prepare_output(output)
        directories = self._create_layout(output)
        challenges = self._selected_challenges()
        overview_rows = []
        image_rows = []
        emitted = []

        docker = None if self.options.skip_docker else shutil.which("docker")
        has_containers = any(self._is_containerized(metadata) for _, metadata in challenges)
        if has_containers and not self.options.skip_docker and not docker:
            message = "docker CLI unavailable; Docker tar export skipped"
            if self.options.require_docker:
                raise PackingError(message)
            self.warnings.append(message)

        for challenge_dir, metadata in challenges:
            record = self._pack_challenge(
                challenge_dir, metadata, directories, docker
            )
            overview_rows.append(self._overview_row(metadata))
            if record["image_row"]:
                image_rows.append(record["image_row"])
            emitted.append(record)

        self._write_workbook(
            directories["题库资源"] / "ctf-overview.xlsx",
            OVERVIEW_HEADERS,
            overview_rows,
        )
        self._write_workbook(
            directories["虚拟机资源"] / "镜像模板.xlsx",
            IMAGE_HEADERS,
            image_rows,
        )
        return {
            "output": str(output),
            "challenges": len(challenges),
            "emitted": emitted,
            "warnings": self.warnings,
            "errors": self.errors,
        }

    @staticmethod
    def _prepare_output(output: Path) -> None:
        output.mkdir(parents=True, exist_ok=True)
        for owned_name in ("工具", "题库资源", "虚拟机资源"):
            owned_path = output / owned_name
            if owned_path.is_symlink():
                raise PackingError(f"refusing symlinked output path: {owned_path}")
            if owned_path.exists():
                if not owned_path.is_dir():
                    raise PackingError(f"output path is not a directory: {owned_path}")
                shutil.rmtree(owned_path)

    def _create_layout(self, output: Path) -> dict[str, Path]:
        paths = {
            "工具": output / "工具",
            "题库资源": output / "题库资源",
            "deploy": output / "题库资源" / "deploy",
            "enclosure": output / "题库资源" / "deploy" / "enclosure",
            "report": output / "题库资源" / "deploy" / "report",
            "虚拟机资源": output / "虚拟机资源",
            "docker-tar": output / "虚拟机资源" / "docker-tar",
        }
        for path in paths.values():
            path.mkdir(parents=True, exist_ok=True)
        return paths

    def _selected_challenges(self) -> list[tuple[Path, dict[str, Any]]]:
        selected = []
        for metadata_path in sorted(self.paths.challenges.glob("*/*/metadata.json")):
            metadata = read_json(metadata_path, {})
            if not isinstance(metadata, dict) or metadata.get("build_status") != "passed":
                continue
            selected.append((metadata_path.parent, metadata))
        return selected

    def _pack_challenge(
        self,
        challenge_dir: Path,
        metadata: dict[str, Any],
        directories: dict[str, Path],
        docker: str | None,
    ) -> dict[str, Any]:
        challenge_id = str(metadata.get("id") or challenge_dir.name)
        category = str(metadata.get("category") or challenge_dir.parent.name).lower()
        prefix = CATEGORY_PREFIXES.get(category, category)
        delivery_name = self._safe_name(str(metadata.get("delivery_name") or challenge_id))
        stem = f"js-{prefix}-{delivery_name}"

        tools_zip = directories["工具"] / f"{stem}exp.zip"
        self._write_tools_zip(challenge_dir, tools_zip)

        pdf_path = directories["report"] / f"{stem}.pdf"
        self._render_pdf(challenge_dir / "writeup" / "wp.md", pdf_path)

        deploy_zip = None
        if self._is_containerized(metadata):
            if metadata.get("port") in (None, ""):
                raise PackingError(f"{challenge_id}: containerized challenge has no port")
            deploy_zip = directories["deploy"] / f"{stem}.zip"
            deploy_dir = challenge_dir / "deploy"
            if not deploy_dir.is_dir():
                raise PackingError(f"{challenge_id}: missing deploy directory")
            self._write_zip(deploy_zip, self._tree_members(deploy_dir, Path("deploy")))

        enclosure_zip = None
        if self._should_emit_enclosure(category):
            members = list(self._enclosure_members(challenge_dir))
            if not members:
                raise PackingError(f"{challenge_id}: required enclosure is empty")
            enclosure_zip = directories["enclosure"] / f"{stem}.zip"
            self._write_zip(enclosure_zip, members)

        tar_path = None
        image_row = None
        if deploy_zip and docker:
            tar_path, image_row = self._save_docker(
                docker, metadata, delivery_name, directories["docker-tar"]
            )

        return {
            "id": challenge_id,
            "tools": str(tools_zip),
            "report": str(pdf_path),
            "deploy": str(deploy_zip) if deploy_zip else None,
            "enclosure": str(enclosure_zip) if enclosure_zip else None,
            "docker_tar": str(tar_path) if tar_path else None,
            "image_row": image_row,
        }

    def _write_tools_zip(self, challenge_dir: Path, destination: Path) -> None:
        writeup = challenge_dir / "writeup" / "wp.md"
        solve = challenge_dir / "solve"
        if not writeup.is_file():
            raise PackingError(f"{challenge_dir.name}: missing writeup/wp.md")
        if not solve.is_dir():
            raise PackingError(f"{challenge_dir.name}: missing solve directory")

        members = [(writeup, Path("wp.md"))]
        solver_files = [
            path
            for path in sorted(solve.rglob("*"))
            if path.is_file() and not path.is_symlink()
        ]
        if not solver_files:
            raise PackingError(f"{challenge_dir.name}: solve directory is empty")
        for path in solver_files:
            relative = path.relative_to(solve)
            if relative == Path("solve.py"):
                relative = Path("exp.py")
            members.append((path, relative))
        self._write_zip(destination, members)

    def _render_pdf(self, markdown_path: Path, destination: Path) -> None:
        source = markdown_path.read_text(encoding="utf-8")
        if not re.search(r"[\u3400-\u9fff]", source):
            self.warnings.append(f"{markdown_path}: writeup contains no CJK text")
        try:
            from reportlab.lib.enums import TA_LEFT
            from reportlab.lib.pagesizes import A4
            from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
            from reportlab.lib.units import mm
            from reportlab.pdfbase import pdfmetrics
            from reportlab.pdfbase.cidfonts import UnicodeCIDFont
            from reportlab.platypus import (
                Paragraph,
                Preformatted,
                SimpleDocTemplate,
                Spacer,
            )
        except ImportError as exc:
            raise PackingError(
                "PDF dependencies unavailable; run `uv sync`"
            ) from exc

        pdfmetrics.registerFont(UnicodeCIDFont("STSong-Light"))
        styles = getSampleStyleSheet()
        body_style = ParagraphStyle(
            "ChineseBody",
            parent=styles["BodyText"],
            fontName="STSong-Light",
            fontSize=10.5,
            leading=17,
            alignment=TA_LEFT,
            spaceAfter=7,
        )
        heading_styles = {
            level: ParagraphStyle(
                f"ChineseHeading{level}",
                parent=styles[f"Heading{min(level, 3)}"],
                fontName="STSong-Light",
                fontSize={1: 20, 2: 16, 3: 13}.get(level, 11),
                leading={1: 26, 2: 21, 3: 18}.get(level, 16),
                spaceBefore=10,
                spaceAfter=8,
            )
            for level in range(1, 7)
        }
        code_style = ParagraphStyle(
            "Code",
            parent=styles["Code"],
            fontName="Courier",
            fontSize=8.5,
            leading=11,
            leftIndent=6,
            rightIndent=6,
            spaceBefore=4,
            spaceAfter=8,
        )
        story = []
        paragraph: list[str] = []
        code: list[str] = []
        in_code = False

        def flush_paragraph() -> None:
            if paragraph:
                text = " ".join(line.strip() for line in paragraph)
                story.append(Paragraph(self._escape_pdf_text(text), body_style))
                paragraph.clear()

        for line in source.splitlines():
            if line.startswith("```"):
                if in_code:
                    story.append(Preformatted("\n".join(code), code_style))
                    code.clear()
                else:
                    flush_paragraph()
                in_code = not in_code
                continue
            if in_code:
                code.append(line)
                continue
            heading = re.match(r"^(#{1,6})\s+(.*)$", line)
            if heading:
                flush_paragraph()
                level = len(heading.group(1))
                story.append(
                    Paragraph(
                        self._escape_pdf_text(heading.group(2)),
                        heading_styles[level],
                    )
                )
            elif not line.strip():
                flush_paragraph()
                story.append(Spacer(1, 2 * mm))
            else:
                paragraph.append(line)
        flush_paragraph()
        if code:
            story.append(Preformatted("\n".join(code), code_style))

        destination.parent.mkdir(parents=True, exist_ok=True)
        document = SimpleDocTemplate(
            str(destination),
            pagesize=A4,
            rightMargin=18 * mm,
            leftMargin=18 * mm,
            topMargin=18 * mm,
            bottomMargin=18 * mm,
            title=markdown_path.stem,
            author="Challenge Factory",
        )
        document.build(story)
        if not destination.read_bytes().startswith(b"%PDF"):
            raise PackingError(f"{markdown_path}: PDF renderer produced invalid output")

    def _write_workbook(
        self, destination: Path, headers: list[str], rows: list[list[Any]]
    ) -> None:
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill
        except ImportError as exc:
            raise PackingError(
                "XLSX dependency unavailable; run `uv sync`"
            ) from exc

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "清单"
        sheet.append(headers)
        for row in rows:
            sheet.append(row)
        fill = PatternFill("solid", fgColor="D9EAF7")
        for cell in sheet[1]:
            cell.font = Font(bold=True)
            cell.fill = fill
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for column in sheet.columns:
            width = max(len(str(cell.value or "")) for cell in column) + 2
            sheet.column_dimensions[column[0].column_letter].width = min(width, 60)
        workbook.save(destination)

    def _overview_row(self, metadata: dict[str, Any]) -> list[Any]:
        category = str(metadata.get("category", "")).lower()
        description = metadata.get("description") or metadata.get("learning_objective") or ""
        technique = metadata.get("primary_technique") or metadata.get("technique") or ""
        return [
            metadata.get("id", ""),
            metadata.get("delivery_name") or metadata.get("title", ""),
            description,
            CATEGORY_LABELS.get(category, category.title()),
            str(metadata.get("difficulty", "")).title(),
            technique,
            metadata.get("points", ""),
            metadata.get("flag_format") or "flag{...}",
            "通过",
        ]

    def _save_docker(
        self,
        docker: str,
        metadata: dict[str, Any],
        delivery_name: str,
        output: Path,
    ) -> tuple[Path | None, list[Any] | None]:
        generated_on = self.options.generated_on or date.today()
        port = metadata.get("port", "")
        tar_name = f"{delivery_name}[{port}]-{generated_on:%Y%m%d}.tar"
        tar_path = output / tar_name
        image = metadata.get("docker_image") or f"{metadata.get('id')}:{generated_on:%Y%m}"
        process = subprocess.run(
            [docker, "save", "-o", str(tar_path), str(image)],
            text=True,
            capture_output=True,
            check=False,
        )
        if process.returncode != 0:
            message = (
                f"{metadata.get('id')}: docker save failed for {image}: "
                f"{process.stderr.strip() or process.stdout.strip()}"
            )
            if self.options.require_docker:
                raise PackingError(message)
            self.errors.append(message)
            tar_path.unlink(missing_ok=True)
            return None, None
        return tar_path, [
            metadata.get("delivery_name") or metadata.get("title") or metadata.get("id"),
            tar_name,
            port,
            metadata.get("base_image", ""),
            metadata.get("start_command", ""),
        ]

    def _should_emit_enclosure(self, category: str) -> bool:
        rule = ENCLOSURE_RULES.get(category, "required")
        if rule == "skip":
            return False
        if category == "pwn":
            return self.options.include_pwn_attachments
        return rule == "required"

    @staticmethod
    def _is_containerized(metadata: dict[str, Any]) -> bool:
        category = str(metadata.get("category", "")).lower()
        deployment = str(metadata.get("deployment", "")).lower()
        return category in {"web", "pwn"} or "docker" in deployment or bool(
            metadata.get("docker_image")
        )

    @staticmethod
    def _safe_name(value: str) -> str:
        normalized = re.sub(r"[^A-Za-z0-9._-]+", "_", value).strip("._-")
        if not normalized:
            raise PackingError(f"invalid delivery name: {value!r}")
        return normalized

    @staticmethod
    def _escape_pdf_text(value: str) -> str:
        return (
            value.replace("&", "&amp;")
            .replace("<", "&lt;")
            .replace(">", "&gt;")
        )

    @staticmethod
    def _tree_members(root: Path, archive_root: Path) -> Iterable[tuple[Path, Path]]:
        for path in sorted(root.rglob("*")):
            if path.is_file() and not path.is_symlink():
                yield path, archive_root / path.relative_to(root)

    def _enclosure_members(self, challenge_dir: Path) -> Iterable[tuple[Path, Path]]:
        for directory_name in ("dist", "attachments"):
            root = challenge_dir / directory_name
            if root.is_dir():
                yield from self._tree_members(root, Path())

    @staticmethod
    def _write_zip(
        destination: Path, members: Iterable[tuple[Path, Path]]
    ) -> None:
        destination.parent.mkdir(parents=True, exist_ok=True)
        with zipfile.ZipFile(
            destination, "w", compression=zipfile.ZIP_DEFLATED
        ) as archive:
            for source, archive_path in members:
                info = zipfile.ZipInfo.from_file(source, archive_path.as_posix())
                info.date_time = (1980, 1, 1, 0, 0, 0)
                info.compress_type = zipfile.ZIP_DEFLATED
                archive.writestr(info, source.read_bytes())

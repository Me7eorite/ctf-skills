"""Workbook inventory helpers."""

from __future__ import annotations

from pathlib import Path
from typing import Any

from packing.errors import PackingError

CATEGORY_LABELS = {
    "crypto": "Crypto",
    "web": "Web",
    "pwn": "Pwn",
    "re": "Reverse",
    "reverse": "Reverse",
    "misc": "Misc",
    "stego": "Stego",
    "forensics": "Forensics",
    "ics": "ICS",
    "ai": "AI",
    "cloud": "Cloud",
    "mobile": "Mobile",
    "blockchain": "Blockchain",
    "iot": "IoT",
    "auto": "Auto",
    "data": "Data",
    "malware": "Malware",
    "osint": "OSINT",
}


def _write_workbook(destination: Path, headers: list[str], rows: list[list[Any]]) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
    except ImportError as exc:
        raise PackingError("XLSX dependency unavailable; run `uv sync`") from exc

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "清单"
    sheet.append(headers)
    for row in rows:
        sheet.append(row)
    fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = fill
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for column in sheet.columns:
        width = max(len(str(cell.value or "")) for cell in column) + 2
        sheet.column_dimensions[column[0].column_letter].width = min(width, 60)
    workbook.save(destination)


def _overview_row(metadata: dict[str, Any]) -> list[Any]:
    category = str(metadata.get("category", "")).lower()
    description = metadata.get("description") or metadata.get("learning_objective") or ""
    technique = metadata.get("primary_technique") or metadata.get("technique") or ""
    return [
        metadata.get("id", ""),
        metadata.get("delivery_name") or metadata.get("title", ""),
        description,
        CATEGORY_LABELS.get(category, category.title()),
        str(metadata.get("difficulty", "")).title(),
        technique,
        metadata.get("points", ""),
        metadata.get("flag_format") or "flag{...}",
        "通过",
    ]
